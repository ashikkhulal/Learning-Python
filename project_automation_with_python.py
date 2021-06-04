# in this project, we will try to automate reading the data from the excel file and do something with information by
# writing a logic and we will use an external package for excel to work with it


import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)
    inventory_price_header = product_list.cell(1, 5)

    # calculation of number of products per supplier
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        products_per_supplier[supplier_name] = 1

    # calculation of total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # logic products with inventory less than 10
    if inventory < 10:
        products_under_10_inv[int(product_num)] = int(inventory)

    # add value for total inventory price
    inventory_price.value = inventory * price
    inventory_price_header.value = "Inventory * Price"

inv_file.save("inventory_solution.xlsx")
print("inventory_solution.xlsx file has been created and saved...")
print("\nSummary:")
print(f"Company and total inventories: {products_per_supplier}")
print(f"Companies and total prices: {total_value_per_supplier}")
print(f"These products are low in stock (product number: inventory): {products_under_10_inv}")